import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re

# ═══════════════════════════════════════════════════════════════
#  PDF 解析：進項憑證明細表
# ═══════════════════════════════════════════════════════════════
def parse_jinxiang_pdf(f):
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("需要安裝 pdfplumber：pip install pdfplumber")
    f.seek(0)
    with pdfplumber.open(f) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    records = []
    cat_re  = re.compile(r'憑證類別[:：]\s*(.+?)\s+總冊數')
    skip_re = re.compile(r'^(進項憑證|用戶|統一編號|序號|送件年月|合計|固\s|總計|頁\s次)')
    current_cat = ""
    pending_sub = 0
    for line in text.splitlines():
        line = line.strip()
        if not line or skip_re.match(line):
            continue
        m = cat_re.search(line)
        if m:
            current_cat = m.group(1).strip()
            pending_sub = 0
            continue
        parts = line.split()
        # 子明細（冊型展開）
        if pending_sub > 0 and re.match(r'^\d{2}$', parts[0]) and (len(parts) < 2 or parts[1] != '進貨'):
            try:
                records.append({
                    '發票號碼': parts[4],
                    '申報金額': int(parts[1].replace(',', '')),
                    '申報稅額': int(parts[2].replace(',', '')),
                    '申報日期': parts[3],
                    '賣方統編': parts[5] if len(parts) > 5 else '',
                    '供商名稱': ' '.join(parts[6:]) if len(parts) > 6 else '',
                    '憑證類別': current_cat,
                    '扣抵': 'Y'
                })
                pending_sub -= 1
                continue
            except Exception:
                pass
        # 主明細：NN 進貨 金額 5* 稅額 日期 發票號碼 [+N] [統編] [名稱] [Y/N]
        if len(parts) >= 7 and parts[1] == '進貨':
            try:
                amount = int(parts[2].replace(',', ''))
                tax    = int(parts[4].replace(',', ''))
                date   = parts[5]
                inv_no = parts[6]
                i = 7; n_sub = 0
                if i < len(parts) and parts[i].startswith('+'):
                    n_sub = int(parts[i][1:]); i += 1
                seller = ''
                if i < len(parts) and re.match(r'^\d{8}$', parts[i]):
                    seller = parts[i]; i += 1
                ded  = parts[-1] if parts[-1] in ('Y', 'N') else 'Y'
                name = ' '.join(parts[i:-1]) if parts[-1] in ('Y', 'N') else ' '.join(parts[i:])
                if n_sub > 0:
                    pending_sub = n_sub
                else:
                    records.append({
                        '發票號碼': inv_no, '申報金額': amount, '申報稅額': tax,
                        '申報日期': date, '賣方統編': seller, '供商名稱': name,
                        '憑證類別': current_cat, '扣抵': ded
                    })
            except Exception:
                pass
    cols = ['發票號碼', '申報金額', '申報稅額', '申報日期', '賣方統編', '供商名稱', '憑證類別', '扣抵']
    df = pd.DataFrame(records) if records else pd.DataFrame(columns=cols)
    if not df.empty:
        df['發票號碼'] = df['發票號碼'].astype(str).str.strip()
    return df


# ═══════════════════════════════════════════════════════════════
#  Excel 自動偵測（外帳 gov / 內帳 acc）
# ═══════════════════════════════════════════════════════════════
def detect_and_load(f):
    f.seek(0)
    raw = pd.read_excel(f, sheet_name=0, header=None)
    GOV_KEYS = {"發票狀態", "賣方名稱", "銷售額合計", "買方統一編號"}
    ACC_KEYS = {"交易模式", "收支日期", "會計項目", "資金帳戶"}
    file_type = None; header_row = 0
    for i in range(min(4, len(raw))):
        vals = set(str(v) for v in raw.iloc[i].tolist())
        if len(GOV_KEYS & vals) >= 2: file_type = "gov"; header_row = i; break
        if len(ACC_KEYS & vals) >= 1: file_type = "acc"; header_row = i; break
    if file_type is None:
        for i in range(min(4, len(raw))):
            vals = set(str(v) for v in raw.iloc[i].tolist())
            if "發票狀態" in vals or "賣方名稱" in vals: file_type = "gov"; header_row = i; break
            if "交易模式" in vals or "收支日期" in vals: file_type = "acc"; header_row = i; break
    if file_type is None:
        raise ValueError(f"無法判斷格式，前幾欄：{[str(v) for v in raw.iloc[0].tolist()[:8]]}")
    df = raw.iloc[header_row + 1:].copy()
    df.columns = [str(c) for c in raw.iloc[header_row].tolist()]
    df = df.reset_index(drop=True)
    if "發票號碼" not in df.columns:
        guess = next((c for c in df.columns if "發票" in c and "號" in c), None)
        if guess:
            df = df.rename(columns={guess: "發票號碼"})
        else:
            raise ValueError(f"找不到發票號碼欄位：{list(df.columns)[:10]}")
    df["發票號碼"] = df["發票號碼"].astype(str).str.strip()
    return df, file_type


# ═══════════════════════════════════════════════════════════════
#  共用樣式工具
# ═══════════════════════════════════════════════════════════════
_FN   = "Arial"
_AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _fill(hex_): return PatternFill("solid", start_color=hex_)
def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

C_HDR = _fill("2F4F4F"); C_GOV = _fill("FFD7D7"); C_ACC = _fill("D7EBFF")
C_PDF = _fill("E8D7FF"); C_BTH = _fill("E8F5E9"); C_DIF = _fill("FFF3CD")
C_KCK = _fill("FFE0B2"); C_VOI = _fill("F5F5F5"); C_SCT = _fill("EAEAEA")

EXCL = {"nan", "收據", "無", ""}

def _write_hdr(ws, headers, widths):
    ws.row_dimensions[1].height = 22
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i)
        c.value = h
        c.font  = Font(name=_FN, bold=True, color="FFFFFF", size=11)
        c.fill  = C_HDR; c.alignment = _AL_C; c.border = _bdr()
        ws.column_dimensions[get_column_letter(i)].width = w

def _wc(ws, r, c, v, fill):
    cell = ws.cell(row=r, column=c)
    cell.value = v; cell.fill = fill
    cell.font  = Font(name=_FN, size=10)
    cell.border = _bdr(); cell.alignment = _AL_L

def _sc(row, col, default=""):
    try:
        v = row[col]; return v if pd.notna(v) else default
    except Exception:
        return default

def _date(row, col):
    v = _sc(row, col)
    return str(v)[:10] if v != "" else ""

def _amt(row, col):
    v = _sc(row, col)
    try: return float(v)
    except Exception: return ""

def _summary_sheet(ws, title, rows, legend):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    t = ws["A1"]; t.value = title
    t.font = Font(name=_FN, bold=True, size=14, color="FFFFFF")
    t.fill = C_HDR; t.alignment = _AL_C; ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 46; ws.column_dimensions["B"].width = 14
    r = 3
    for lbl, val in rows:
        if val is None:
            c = ws.cell(row=r, column=1)
            c.value = lbl; c.font = Font(name=_FN, bold=True, size=10); c.fill = C_SCT
        else:
            c1 = ws.cell(row=r, column=1)
            c1.value = lbl; c1.font = Font(name=_FN, size=10)
            fl = legend.get(lbl, PatternFill()); c1.fill = fl
            if lbl in legend: c1.border = _bdr()
            c2 = ws.cell(row=r, column=2)
            c2.value = val; c2.font = Font(name=_FN, bold=True, size=10)
            c2.alignment = Alignment(horizontal="center")
        r += 1


# ═══════════════════════════════════════════════════════════════
#  模式① 報告：申報前核對（電子外帳 vs 內帳）
# ═══════════════════════════════════════════════════════════════
def _build_step1(df_gov, df_acc, only_gov, only_acc, both):
    wb = Workbook()
    if "發票狀態" in df_gov.columns:
        valid_g = only_gov & set(df_gov[df_gov["發票狀態"] == "開立已確認"]["發票號碼"])
        void_g  = only_gov & set(df_gov[df_gov["發票狀態"] == "作廢已確認"]["發票號碼"])
    else:
        valid_g = only_gov; void_g = set()

    ws1 = wb.active; ws1.title = "差異摘要"
    _summary_sheet(ws1, "進項發票比對報告（①申報前核對）— 差異摘要", [
        ("", ""), ("📋 比對統計", None),
        ("電子外帳（財政部）筆數",   len(df_gov)),
        ("內帳（會計軟體）進項筆數", len(df_acc)),
        ("兩邊吻合",               len(both)), ("", ""),
        ("⚠️ 差異狀況", None),
        ("外帳有、內帳沒有（共）",   len(only_gov)),
        ("  └ 開立已確認",          len(valid_g)),
        ("  └ 作廢已確認",          len(void_g)),
        ("內帳有、外帳沒有",         len(only_acc)), ("", ""),
        ("📌 顏色說明", None),
        ("淡紅底", "外帳有、內帳沒有"),
        ("淡藍底", "內帳有、外帳沒有"),
        ("淡綠底", "兩邊吻合（金額一致）"),
        ("淡黃底", "兩邊都有但金額不同"),
        ("淺灰底", "作廢 / 收據"),
    ], {"淡紅底": C_GOV, "淡藍底": C_ACC, "淡綠底": C_BTH, "淡黃底": C_DIF, "淺灰底": C_VOI})

    ws2 = wb.create_sheet("外帳有_內帳沒有"); ws2.sheet_view.showGridLines = False
    _write_hdr(ws2,
        ["發票號碼","發票狀態","發票日期","賣方名稱","銷售額合計","營業稅","總計"],
        [14,12,12,32,12,10,12])
    diff = df_gov[df_gov["發票號碼"].isin(only_gov)].copy()
    if "發票日期" in diff.columns: diff = diff.sort_values("發票日期")
    for rn, (_, row) in enumerate(diff.iterrows(), 2):
        st_ = _sc(row, "發票狀態"); fl = C_VOI if st_ == "作廢已確認" else C_GOV
        for c, v in enumerate([_sc(row,"發票號碼"), st_, _date(row,"發票日期"),
                                _sc(row,"賣方名稱"), _amt(row,"銷售額合計"),
                                _amt(row,"營業稅"), _amt(row,"總計")], 1):
            _wc(ws2, rn, c, v, fl)

    ws3 = wb.create_sheet("內帳有_外帳沒有"); ws3.sheet_view.showGridLines = False
    _write_hdr(ws3,
        ["發票號碼","憑證日期","收支日期","對象","發票金額","稅額","銷售額","附註說明"],
        [14,12,12,24,12,8,12,36])
    diff = df_acc[df_acc["發票號碼"].isin(only_acc)].copy()
    if "憑證日期" in diff.columns: diff = diff.sort_values("憑證日期")
    for rn, (_, row) in enumerate(diff.iterrows(), 2):
        inv = str(_sc(row, "發票號碼")); fl = C_VOI if inv in EXCL else C_ACC
        for c, v in enumerate([inv, _date(row,"憑證日期"), _date(row,"收支日期"),
                                _sc(row,"對象"), _amt(row,"發票金額"),
                                _amt(row,"稅額"), _amt(row,"銷售額"), _sc(row,"附註說明")], 1):
            _wc(ws3, rn, c, v, fl)

    ws4 = wb.create_sheet("兩邊都有_金額核對"); ws4.sheet_view.showGridLines = False
    _write_hdr(ws4,
        ["發票號碼","外帳日期","外帳賣方","外帳金額","內帳憑證日期","內帳對象","內帳金額","差異"],
        [14,12,30,12,13,24,12,10])
    rn = 2
    for inv in sorted(both):
        gr = df_gov[df_gov["發票號碼"] == inv]; ar = df_acc[df_acc["發票號碼"] == inv]
        if gr.empty or ar.empty: continue
        g = gr.iloc[0]; a = ar.iloc[0]
        ga = _amt(g, "總計"); aa = _amt(a, "發票金額")
        dv = (ga - aa) if isinstance(ga, float) and isinstance(aa, float) else ""
        fl = C_DIF if isinstance(dv, float) and abs(dv) > 0.5 else C_BTH
        for c, v in enumerate([inv, _date(g,"發票日期"), _sc(g,"賣方名稱"), ga,
                                _date(a,"憑證日期"), _sc(a,"對象"), aa,
                                dv if isinstance(dv, float) and abs(dv) > 0.5 else ""], 1):
            _wc(ws4, rn, c, v, fl)
        rn += 1
    return wb


# ═══════════════════════════════════════════════════════════════
#  模式② 報告：申報後覆核（申報PDF vs 內帳）
# ═══════════════════════════════════════════════════════════════
def _build_step2(df_pdf, df_acc, only_pdf, only_acc, both):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "差異摘要"
    _summary_sheet(ws1, "進項發票比對報告（②申報後覆核）— 差異摘要", [
        ("", ""), ("📋 比對統計", None),
        ("申報進項明細筆數", len(df_pdf)),
        ("內帳進項筆數",    len(df_acc)),
        ("兩邊吻合",       len(both)), ("", ""),
        ("⚠️ 差異狀況", None),
        ("申報有、內帳沒有", len(only_pdf)),
        ("內帳有、申報沒有", len(only_acc)), ("", ""),
        ("📌 顏色說明", None),
        ("淡紫底", "申報有、內帳沒有"),
        ("淡藍底", "內帳有、申報沒有"),
        ("淡綠底", "兩邊吻合（金額一致）"),
        ("淡黃底", "金額不同"),
    ], {"淡紫底": C_PDF, "淡藍底": C_ACC, "淡綠底": C_BTH, "淡黃底": C_DIF})

    ws2 = wb.create_sheet("申報有_內帳沒有"); ws2.sheet_view.showGridLines = False
    _write_hdr(ws2,
        ["發票號碼","申報日期","賣方統編","供商名稱","申報金額","申報稅額","憑證類別","扣抵"],
        [14,12,12,26,12,10,20,8])
    diff = df_pdf[df_pdf["發票號碼"].isin(only_pdf)].copy()
    if "申報日期" in diff.columns: diff = diff.sort_values("申報日期")
    for rn, (_, row) in enumerate(diff.iterrows(), 2):
        for c, v in enumerate([_sc(row,"發票號碼"), _sc(row,"申報日期"),
                                _sc(row,"賣方統編"), _sc(row,"供商名稱"),
                                _amt(row,"申報金額"), _amt(row,"申報稅額"),
                                _sc(row,"憑證類別"), _sc(row,"扣抵")], 1):
            _wc(ws2, rn, c, v, C_PDF)

    ws3 = wb.create_sheet("內帳有_申報沒有"); ws3.sheet_view.showGridLines = False
    _write_hdr(ws3,
        ["發票號碼","憑證日期","收支日期","對象","發票金額","稅額","銷售額","附註說明"],
        [14,12,12,24,12,8,12,36])
    diff = df_acc[df_acc["發票號碼"].isin(only_acc)].copy()
    if "憑證日期" in diff.columns: diff = diff.sort_values("憑證日期")
    for rn, (_, row) in enumerate(diff.iterrows(), 2):
        inv = str(_sc(row, "發票號碼")); fl = C_VOI if inv in EXCL else C_ACC
        for c, v in enumerate([inv, _date(row,"憑證日期"), _date(row,"收支日期"),
                                _sc(row,"對象"), _amt(row,"發票金額"),
                                _amt(row,"稅額"), _amt(row,"銷售額"), _sc(row,"附註說明")], 1):
            _wc(ws3, rn, c, v, fl)

    ws4 = wb.create_sheet("兩邊都有_金額核對"); ws4.sheet_view.showGridLines = False
    _write_hdr(ws4,
        ["發票號碼","申報日期","供商名稱","申報金額","內帳憑證日期","內帳對象","內帳金額","差異"],
        [14,12,26,12,13,24,12,10])
    rn = 2
    for inv in sorted(both):
        pr = df_pdf[df_pdf["發票號碼"] == inv]; ar = df_acc[df_acc["發票號碼"] == inv]
        if pr.empty or ar.empty: continue
        p = pr.iloc[0]; a = ar.iloc[0]
        pa = _amt(p, "申報金額"); aa = _amt(a, "發票金額")
        dv = (pa - aa) if isinstance(pa, float) and isinstance(aa, float) else ""
        fl = C_DIF if isinstance(dv, float) and abs(dv) > 0.5 else C_BTH
        for c, v in enumerate([inv, _sc(p,"申報日期"), _sc(p,"供商名稱"), pa,
                                _date(a,"憑證日期"), _sc(a,"對象"), aa,
                                dv if isinstance(dv, float) and abs(dv) > 0.5 else ""], 1):
            _wc(ws4, rn, c, v, fl)
        rn += 1
    return wb


# ═══════════════════════════════════════════════════════════════
#  模式③ 報告：三方比對
# ═══════════════════════════════════════════════════════════════
def _build_threeway(df_gov, df_pdf, df_acc):
    wb = Workbook()
    gov_s = set(df_gov["發票號碼"]) - EXCL
    pdf_s = set(df_pdf["發票號碼"]) - EXCL
    acc_s = set(df_acc["發票號碼"]) - EXCL

    all3   = gov_s & pdf_s & acc_s
    kicked = (gov_s & acc_s) - pdf_s   # 被踢除
    miss   = (gov_s & pdf_s) - acc_s   # 漏入帳
    paper  = (pdf_s & acc_s) - gov_s   # 紙本已申報
    solo_a = acc_s - gov_s - pdf_s     # 孤立內帳
    solo_g = gov_s - pdf_s - acc_s     # 孤立電子

    ws1 = wb.active; ws1.title = "三方比對摘要"
    _summary_sheet(ws1, "進項發票三方比對報告 — 摘要", [
        ("", ""), ("📋 三方資料筆數", None),
        ("電子外帳（財政部）筆數",  len(df_gov)),
        ("申報進項明細（PDF）筆數", len(df_pdf)),
        ("內帳（會計軟體）筆數",   len(df_acc)), ("", ""),
        ("📊 比對結果", None),
        ("✅ 三方吻合",                       len(all3)),
        ("🔶 被踢除（電子+內帳有，申報沒有）",  len(kicked)),
        ("✅ 紙本已申報（申報+內帳有，無電子）", len(paper)),
        ("❌ 漏入帳（電子+申報有，內帳沒有）",  len(miss)),
        ("⬜ 孤立內帳（僅內帳有）",             len(solo_a)),
        ("⬜ 孤立電子（僅電子有）",             len(solo_g)), ("", ""),
        ("📌 顏色說明", None),
        ("橘色底", "被踢除（需確認原因）"),
        ("淡紅底", "漏入帳"),
        ("淡綠底", "三方吻合 / 紙本已申報"),
        ("淡藍底", "孤立內帳"),
        ("淺灰底", "孤立電子"),
    ], {"橘色底": C_KCK, "淡紅底": C_GOV, "淡綠底": C_BTH, "淡藍底": C_ACC, "淺灰底": C_VOI})

    # ── 被踢除（最重要）
    ws2 = wb.create_sheet("被踢除"); ws2.sheet_view.showGridLines = False
    _write_hdr(ws2,
        ["發票號碼","電子外帳日期","電子賣方名稱","電子金額","內帳憑證日期","內帳對象","內帳金額"],
        [14,12,26,12,13,24,12])
    rn = 2
    for inv in sorted(kicked):
        gr = df_gov[df_gov["發票號碼"] == inv]; ar = df_acc[df_acc["發票號碼"] == inv]
        g = gr.iloc[0] if not gr.empty else None
        a = ar.iloc[0] if not ar.empty else None
        for c, v in enumerate([inv,
            _date(g,"發票日期") if g is not None else "",
            _sc(g,"賣方名稱") if g is not None else "",
            _amt(g,"總計") if g is not None else "",
            _date(a,"憑證日期") if a is not None else "",
            _sc(a,"對象") if a is not None else "",
            _amt(a,"發票金額") if a is not None else ""], 1):
            _wc(ws2, rn, c, v, C_KCK)
        rn += 1

    # ── 漏入帳
    ws3 = wb.create_sheet("漏入帳"); ws3.sheet_view.showGridLines = False
    _write_hdr(ws3,
        ["發票號碼","電子日期","電子賣方名稱","電子金額","申報日期","申報供商名稱","申報金額","憑證類別"],
        [14,12,26,12,12,24,12,18])
    rn = 2
    for inv in sorted(miss):
        gr = df_gov[df_gov["發票號碼"] == inv]; pr = df_pdf[df_pdf["發票號碼"] == inv]
        g = gr.iloc[0] if not gr.empty else None
        p = pr.iloc[0] if not pr.empty else None
        for c, v in enumerate([inv,
            _date(g,"發票日期") if g is not None else "",
            _sc(g,"賣方名稱") if g is not None else "",
            _amt(g,"總計") if g is not None else "",
            _sc(p,"申報日期") if p is not None else "",
            _sc(p,"供商名稱") if p is not None else "",
            _amt(p,"申報金額") if p is not None else "",
            _sc(p,"憑證類別") if p is not None else ""], 1):
            _wc(ws3, rn, c, v, C_GOV)
        rn += 1

    # ── 紙本已申報
    ws4 = wb.create_sheet("紙本已申報"); ws4.sheet_view.showGridLines = False
    _write_hdr(ws4,
        ["發票號碼","申報日期","供商名稱","申報金額","憑證類別","內帳對象","內帳金額","金額差異"],
        [14,12,24,12,18,24,12,10])
    rn = 2
    for inv in sorted(paper):
        pr = df_pdf[df_pdf["發票號碼"] == inv]; ar = df_acc[df_acc["發票號碼"] == inv]
        p = pr.iloc[0] if not pr.empty else None
        a = ar.iloc[0] if not ar.empty else None
        pa = _amt(p, "申報金額") if p is not None else ""
        aa = _amt(a, "發票金額") if a is not None else ""
        dv = (pa - aa) if isinstance(pa, float) and isinstance(aa, float) else ""
        fl = C_DIF if isinstance(dv, float) and abs(dv) > 0.5 else C_BTH
        for c, v in enumerate([inv,
            _sc(p,"申報日期") if p is not None else "",
            _sc(p,"供商名稱") if p is not None else "", pa,
            _sc(p,"憑證類別") if p is not None else "",
            _sc(a,"對象") if a is not None else "", aa,
            dv if isinstance(dv, float) and abs(dv) > 0.5 else ""], 1):
            _wc(ws4, rn, c, v, fl)
        rn += 1

    # ── 孤立內帳
    ws5 = wb.create_sheet("孤立內帳"); ws5.sheet_view.showGridLines = False
    _write_hdr(ws5,
        ["發票號碼","憑證日期","收支日期","對象","發票金額","稅額","附註說明"],
        [14,12,12,24,12,8,36])
    diff5 = df_acc[df_acc["發票號碼"].isin(solo_a)].copy()
    if "憑證日期" in diff5.columns: diff5 = diff5.sort_values("憑證日期")
    for rn, (_, row) in enumerate(diff5.iterrows(), 2):
        inv = str(_sc(row, "發票號碼")); fl = C_VOI if inv in EXCL else C_ACC
        for c, v in enumerate([inv, _date(row,"憑證日期"), _date(row,"收支日期"),
                                _sc(row,"對象"), _amt(row,"發票金額"),
                                _amt(row,"稅額"), _sc(row,"附註說明")], 1):
            _wc(ws5, rn, c, v, fl)

    return wb


# ═══════════════════════════════════════════════════════════════
#  Streamlit 介面
# ═══════════════════════════════════════════════════════════════
st.set_page_config(page_title="進項發票比對工具", page_icon="🧾", layout="centered")
st.title("🧾 進項發票比對工具")
st.divider()

MODE_1 = "① 申報前核對（電子外帳 vs 內帳）"
MODE_2 = "② 申報後覆核（申報PDF vs 內帳）"
MODE_3 = "③ 三方比對（電子外帳 ＋ 申報PDF ＋ 內帳）"

mode = st.radio("比對模式", [MODE_1, MODE_2, MODE_3], horizontal=True)
st.divider()


# ────────────────────────── 模式① ──────────────────────────────
if mode == MODE_1:
    st.caption("上傳財政部電子外帳 Excel 和會計軟體內帳 Excel，系統自動辨識格式。")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**📁 檔案一**")
        fa = st.file_uploader("上傳第一個xlsx", type=["xlsx"], key="m1a",
                              label_visibility="collapsed")
    with col2:
        st.markdown("**📁 檔案二**")
        fb = st.file_uploader("上傳第二個xlsx", type=["xlsx"], key="m1b",
                              label_visibility="collapsed")

    if fa and fb:
        st.divider()
        if st.button("🔍 開始比對", use_container_width=True, type="primary", key="btn1"):
            with st.spinner("讀取並比對中…"):
                try: dfa, ta = detect_and_load(fa)
                except Exception as e: st.error(f"檔案一讀取失敗：{e}"); st.stop()
                try: dfb, tb = detect_and_load(fb)
                except Exception as e: st.error(f"檔案二讀取失敗：{e}"); st.stop()
                if ta == tb:
                    st.error("兩個檔案格式相同，請確認是否分別上傳外帳和內帳。"); st.stop()
                df_gov, df_acc = (dfa, dfb) if ta == "gov" else (dfb, dfa)
                gn, an = (fa.name, fb.name) if ta == "gov" else (fb.name, fa.name)
                st.success(f"✅ 外帳 → {gn}　內帳 → {an}")
                gs = set(df_gov["發票號碼"]) - EXCL
                as_ = set(df_acc["發票號碼"]) - EXCL
                og = gs - as_; oa = as_ - gs; bt = gs & as_
                c1, c2, c3 = st.columns(3)
                c1.metric("外帳有、內帳沒有", len(og))
                c2.metric("內帳有、外帳沒有", len(oa))
                c3.metric("兩邊吻合", len(bt))
                try: wb = _build_step1(df_gov, df_acc, og, oa, bt)
                except Exception as e: st.error(f"報告產生失敗：{e}"); st.stop()
                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("⬇️ 下載申報前核對報告", data=buf,
                file_name="進項發票_申報前核對.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


# ────────────────────────── 模式② ──────────────────────────────
elif mode == MODE_2:
    st.caption("上傳會計師申報進項明細 PDF 和會計軟體內帳 Excel，核對兩者是否吻合。")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**📄 申報進項明細 PDF**")
        fpdf = st.file_uploader("上傳PDF", type=["pdf"], key="m2p",
                                label_visibility="collapsed")
    with col2:
        st.markdown("**📁 內帳 Excel**")
        facc = st.file_uploader("上傳內帳xlsx", type=["xlsx"], key="m2a",
                                label_visibility="collapsed")

    if fpdf and facc:
        st.divider()
        if st.button("🔍 開始覆核", use_container_width=True, type="primary", key="btn2"):
            with st.spinner("解析PDF並比對中…"):
                try: df_pdf = parse_jinxiang_pdf(fpdf)
                except Exception as e: st.error(f"PDF解析失敗：{e}"); st.stop()
                try: df_acc, t = detect_and_load(facc)
                except Exception as e: st.error(f"內帳讀取失敗：{e}"); st.stop()
                if t != "acc": st.warning("⚠️ 上傳的Excel看起來不像內帳格式，請確認。")
                st.success(f"✅ 申報PDF {len(df_pdf)} 筆　內帳 {len(df_acc)} 筆")
                ps = set(df_pdf["發票號碼"]) - EXCL
                as_ = set(df_acc["發票號碼"]) - EXCL
                op = ps - as_; oa = as_ - ps; bt = ps & as_
                c1, c2, c3 = st.columns(3)
                c1.metric("申報有、內帳沒有", len(op))
                c2.metric("內帳有、申報沒有", len(oa))
                c3.metric("兩邊吻合", len(bt))
                try: wb = _build_step2(df_pdf, df_acc, op, oa, bt)
                except Exception as e: st.error(f"報告產生失敗：{e}"); st.stop()
                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("⬇️ 下載申報後覆核報告", data=buf,
                file_name="進項發票_申報後覆核.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


# ────────────────────────── 模式③ ──────────────────────────────
else:
    st.caption("同時上傳三份：財政部電子外帳、申報進項明細 PDF、會計軟體內帳，進行完整三方比對。")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("**📁 電子外帳 Excel**")
        fg = st.file_uploader("外帳xlsx", type=["xlsx"], key="m3g",
                              label_visibility="collapsed")
    with col2:
        st.markdown("**📄 申報進項明細 PDF**")
        fp = st.file_uploader("申報pdf", type=["pdf"], key="m3p",
                              label_visibility="collapsed")
    with col3:
        st.markdown("**📁 內帳 Excel**")
        fa2 = st.file_uploader("內帳xlsx", type=["xlsx"], key="m3a",
                               label_visibility="collapsed")

    if fg and fp and fa2:
        st.divider()
        if st.button("🔍 開始三方比對", use_container_width=True, type="primary", key="btn3"):
            with st.spinner("解析並三方比對中…"):
                try: df_gov, _ = detect_and_load(fg)
                except Exception as e: st.error(f"電子外帳讀取失敗：{e}"); st.stop()
                try: df_pdf = parse_jinxiang_pdf(fp)
                except Exception as e: st.error(f"PDF解析失敗：{e}"); st.stop()
                try: df_acc, _ = detect_and_load(fa2)
                except Exception as e: st.error(f"內帳讀取失敗：{e}"); st.stop()
                st.success(f"✅ 電子外帳 {len(df_gov)} 筆　申報PDF {len(df_pdf)} 筆　內帳 {len(df_acc)} 筆")

                gov_s = set(df_gov["發票號碼"]) - EXCL
                pdf_s = set(df_pdf["發票號碼"]) - EXCL
                acc_s = set(df_acc["發票號碼"]) - EXCL
                all3   = gov_s & pdf_s & acc_s
                kicked = (gov_s & acc_s) - pdf_s
                miss   = (gov_s & pdf_s) - acc_s
                paper  = (pdf_s & acc_s) - gov_s

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("✅ 三方吻合",   len(all3))
                c2.metric("🔶 被踢除",     len(kicked))
                c3.metric("❌ 漏入帳",     len(miss))
                c4.metric("📄 紙本已申報", len(paper))

                try: wb = _build_threeway(df_gov, df_pdf, df_acc)
                except Exception as e: st.error(f"報告產生失敗：{e}"); st.stop()
                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("⬇️ 下載三方比對報告", data=buf,
                file_name="進項發票_三方比對.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
